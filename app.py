from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file
from calculations import calculate_profile, calculate_four_interval_profile, calculate_j_shaped_profile, calculate_s_shaped_profile
from horizontal_calculations import calc_two_int, calc_three_int, calc_four_int_tangential, calc_five_int, calc_four_int
from database import (
    init_db, 
    save_calculation_results, 
    get_calculation_results, 
    get_all_calculations, 
    delete_calculation,
    clear_database
)
from openpyxl import Workbook
from datetime import datetime
import os
from math import pi, sqrt, sin, cos, atan, radians, degrees, tan
import json

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'  # Необходимо для использования session

# Инициализируем базу данных при запуске
init_db()

@app.route('/')
def index():
    return render_template('well_type_selection.html')

@app.route('/directional_well')
def directional_well():
    return render_template('directional_well.html')

@app.route('/history')
def history():
    calculations = get_all_calculations()
    return render_template('history.html', calculations=calculations)

@app.route('/results')
def results():
    if 'calculation_results' not in session:
        return redirect(url_for('calculation'))
    results = session['calculation_results']
    # Устанавливаем profile_type в 0 для горизонтальных скважин, если well_type == 2
    if results.get('well_type') == 2:
        results['profile_type'] = 0
    return render_template('results.html', results=results)

@app.route('/calculate', methods=['POST'])
def calculate():
    try:
        print("Form data:", dict(request.form))  # Отладочный вывод
        data = {
            'H': float(request.form['H']),
            'A': float(request.form['A']),
            'Hv': float(request.form['Hv']),
            'R1': float(request.form['R1']),
            'calculation_type': request.form.get('calculation_type', 'angle_stabilization'),
            'H_end': float(request.form.get('H_end', 0)) or None
        }
        profile_type = int(request.form.get('profile_type', 1))
        print("Profile type:", profile_type)  # Отладочный вывод
        if profile_type in [3, 4]:  # J-образный или S-образный
            data.update({
                'initial_angle': float(request.form['initial_angle']),
                'R4': float(request.form['R4'])
            })
            if profile_type == 4:
                results = calculate_j_shaped_profile(**data)
            else:
                results = calculate_s_shaped_profile(**data)
        elif profile_type == 2:  # Четырехинтервальный
            data.update({
                'initial_angle': float(request.form['initial_angle']),
                'R2': float(request.form['R2'])
            })
            results = calculate_four_interval_profile(**data)
        else:  # Трехинтервальный
            results = calculate_profile(**data)
        if 'error' not in results:
            calculation_id = save_calculation_results(
                depth=data['H'],
                well_type=1,  # наклонная скважина
                profile_type=profile_type,
                type_ha=0,    # всегда 0 для наклонной
                type_h=0,     # всегда 0 для наклонной
                results=results['table_data']
            )
            results['calculation_id'] = calculation_id
            results['profile_type'] = profile_type
            # Добавляем well_type в результаты для сессии
            results['well_type'] = 1
        session['calculation_results'] = results
    except ValueError as e:
        print(f"Ошибка в введенных данных: {str(e)}")
        results = {'error': f"Ошибка в введенных данных: {str(e)}"}
        session['calculation_results'] = results
    except Exception as e:
        print(f"Ошибка при расчете: {str(e)}")
        results = {'error': f"Ошибка при выполнении расчета: {str(e)}"}
        session['calculation_results'] = results
    return redirect(url_for('results'))

@app.route('/get_calculation_data/<int:calculation_id>')
def get_calculation_data(calculation_id):
    try:
        data = get_calculation_results(calculation_id)
        if data is None:
            return jsonify({'error': 'Расчет не найден'})
        return jsonify({
            'headers': [
                'Номер участка',
                'Глубина по вертикали, м',
                'Длина ствола, м',
                'Длина интервала, м',
                'Смещение, м',
                'Зенитный угол, град.',
                'Интенсивность искривления, град./10м'
            ],
            'rows': data['results'].get('rows', []),
            'input_data': {
                'H': data['depth'],
                'profile_type': data['profile_type'],
                'timestamp': data['timestamp'],
                'type_h': data['type_h']
            }
        })
    except Exception as e:
        print(f"Ошибка при получении данных расчета: {str(e)}")
        return jsonify({'error': str(e)})

@app.route('/delete_calculation/<int:calculation_id>', methods=['POST', 'DELETE'])
def delete_calculation_route(calculation_id):
    try:
        delete_calculation(calculation_id)
        return jsonify({'success': True, 'message': 'Расчет успешно удален'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/download_excel/<int:calculation_id>')
def download_excel(calculation_id):
    try:
        data = get_calculation_results(calculation_id)
        if data is None:
            return jsonify({'error': 'Расчет не найден'}), 404
        wb = Workbook()
        ws = wb.active
        ws.title = "Результаты расчета"
        headers = [
            'Номер участка',
            'Глубина по вертикали, м',
            'Длина ствола, м',
            'Длина интервала, м',
            'Смещение, м',
            'Зенитный угол, град.',
            'Интенсивность искривления, град./10м'
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Изменяем цикл, чтобы перебирать все строки в результатах
        rows = data['results']['rows']
        num_rows = len(rows)
        is_directional_well = data.get('type') == 1
        is_horizontal_well = data.get('type') == 2
        type_h = data.get('type_h', 0)

        for row_index, row_data in enumerate(rows):
            current_excel_row = row_index + 2 # Excel rows are 1-indexed and headers are in row 1

            ws.cell(row=current_excel_row, column=1, value=row_data['section'])
            ws.cell(row=current_excel_row, column=2, value=row_data['vertical_depth'])
            ws.cell(row=current_excel_row, column=3, value=row_data['wellbore_length'])
            ws.cell(row=current_excel_row, column=4, value=row_data['interval_length'])
            ws.cell(row=current_excel_row, column=5, value=row_data['displacement'])
            ws.cell(row=current_excel_row, column=6, value=row_data['zenith_angle'])
            ws.cell(row=current_excel_row, column=7, value=row_data['curvature_rate'])

        # After writing all data rows, check if a label row needs to be inserted
        label_row = -1
        label_text = ''

        if is_directional_well:
             # For directional wells, insert label after the last data row
             label_row = num_rows + 2 - 1
             label_text = 'Участок ниже проектной глубины'
        elif is_horizontal_well:
            # Волнообразный тип: метка перед горизонтальным участком (перед последними 3 строками данных)
            if type_h == 2:
                label_row = num_rows + 2 - 3 # Insert before the last 3 data rows
                label_text = 'Горизонтальный участок'
            # Остальные горизонтальные типы: метка перед последней строкой
            elif type_h != 2:
                label_row = num_rows + 2 - 1 # Insert before the last data row
                label_text = 'Горизонтальный участок'

        # If a label needs to be inserted, set the value and merge cells
        if label_row != -1:
            # Insert a new row for the label
            ws.insert_rows(label_row)
            # Set the value of the cell in the newly inserted row
            ws.cell(row=label_row, column=1, value=label_text)
            # Merge the cells in the newly inserted row
            ws.merge_cells(start_row=label_row, start_column=1, end_row=label_row, end_column=len(headers))
            # Apply formatting
            label_cell = ws.cell(row=label_row, column=1)
            from openpyxl.styles import Font, Alignment
            label_cell.font = Font(bold=True)
            label_cell.alignment = Alignment(horizontal='center')

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 20
        filename = f"calculation_{calculation_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join('temp', filename)
        if not os.path.exists('temp'):
            os.makedirs('temp')
        wb.save(filepath)
        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        print(f"Ошибка при создании Excel файла: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/clear_database', methods=['POST'])
def clear_database_route():
    try:
        clear_database()
        return jsonify({'success': True, 'message': 'База данных успешно очищена'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/horizontal_well')
def horizontal_well():
    return render_template('horizontal_well.html')

@app.route('/calculate_horizontal', methods=['POST'])
def calculate_horizontal():
    try:
        print("=== Starting horizontal well calculation ===")
        print("Form data:", dict(request.form))
        print("Request method:", request.method)
        print("Content type:", request.content_type)
        
        data = {
            'H_h': float(request.form['depth']),
            'A_h': float(request.form['displacement']),
            'ang_h': float(request.form['angle']),
            'type_ha': int(request.form.get('type_ha', 1)),
            'type_h': int(request.form.get('type_h', 1))
        }
        print("Initial data:", data)
        
        if data['type_ha'] == 2:
            data['radius_curve_2'] = float(request.form.get('radius_curve_2', 0))
            data['zenith_angle'] = float(request.form.get('zenith_angle', 0))  
        if data['type_ha'] == 3:
            data['radius_curve_4'] = float(request.form.get('radius_curve_4', 0))
            data['radius_curve_2'] = float(request.form.get('radius_curve_2', 0))
            data['zenith_angle'] = float(request.form.get('zenith_angle', 0))  
        if data['type_ha'] == 4:
            data['radius_curve_5'] = float(request.form.get('radius_curve_5', 0))
            data['radius_curve_4'] = float(request.form.get('radius_curve_4', 0))
            data['radius_curve_2'] = float(request.form.get('radius_curve_2', 0))
            data['zenith_angle'] = float(request.form.get('zenith_angle', 0))  
            data['angle_4'] = float(request.form.get('angle_4', 0))
        if data['type_ha'] == 5:
            data['radius_curve_3'] = float(request.form.get('radius_curve_3', 0))
            data['radius_curve_2'] = float(request.form.get('radius_curve_2', 0))
            data['zenith_angle'] = float(request.form.get('zenith_angle', 0))  
            data['angle_3'] = float(request.form.get('angle_3', 0))
        
        # Add length_hor only if type_h is 0
        if data['type_h'] == 0:
            data['length_hor'] = float(request.form.get('length_hor', 0))
            data['deviation_up'] = 0
            data['radius_curve'] = 0
            data['deviation_down'] = 0
        elif data['type_h'] == -1:
            data['length_hor'] = float(request.form.get('length_hor', 0))
            data['deviation_down'] = float(request.form.get('deviation_down', 0))
            data['deviation_up'] = 0
            data['radius_curve'] = 0
        elif data['type_h'] == 1:
            data['deviation_up'] = float(request.form.get('deviation_up', 0))
            data['length_hor'] = float(request.form.get('length_hor', 0))
            data['deviation_down'] = 0
            data['radius_curve'] = 0
        elif data['type_h'] == 2:
            data['radius_curve'] = float(request.form.get('radius_curve', 0))
            data['deviation_up'] = float(request.form.get('deviation_up', 0))
            data['deviation_down'] = float(request.form.get('deviation_down', 0))
            data['length_hor'] = float(request.form.get('length_hor', 0))
        
        # Выполняем расчет в зависимости от типа горизонтального аппарата
        if data['type_ha'] == 1:  # двухинтервальный
            results = calc_two_int(
                data['H_h'], data['A_h'], data['ang_h'],
                data['length_hor'], data['type_ha'], data['type_h'], data['deviation_down'], data['deviation_up'], data['radius_curve']
            )
        elif data['type_ha'] == 2:  # трехинтервальный
            results = calc_three_int(
                data['H_h'], data['A_h'], data['ang_h'],
                data['length_hor'], data['radius_curve_2'], data['zenith_angle'], data['type_ha'], data['type_h'], data['deviation_down'], data['deviation_up'], data['radius_curve']
            )
        elif data['type_ha'] == 3:  # четырехинтервальный
            results = calc_four_int_tangential(
                data['H_h'], data['A_h'], data['ang_h'],
                data['length_hor'], data['radius_curve_2'], data['radius_curve_4'], data['zenith_angle'], data['type_ha'], data['type_h'], data['deviation_down'], data['deviation_up'], data['radius_curve']
            )
        elif data['type_ha'] == 4:  # пятихинтервальный
            results = calc_five_int(
                data['H_h'], data['A_h'], data['ang_h'],
                data['length_hor'], data['radius_curve_2'], data['radius_curve_4'], data['radius_curve_5'], data['zenith_angle'], data['angle_4'], data['type_ha'], data['type_h'], data['deviation_down'], data['deviation_up'], data['radius_curve']
            )
        elif data['type_ha'] == 5:  # 4интервальный
            print("124")
            results = calc_four_int(
                data['H_h'], data['A_h'], data['ang_h'],
                data['length_hor'], data['radius_curve_2'],  data['radius_curve_3'], data['zenith_angle'],  data['angle_3'], data['type_ha'], data['type_h'], data['deviation_down'], data['deviation_up'], data['radius_curve']
            )
        else:
            return render_template('error.html', error="Выбранный тип профиля пока не реализован")

        if 'error' not in results:
            # Сохраняем результаты в базу данных
            calculation_id = save_calculation_results(
                depth=data['H_h'],
                well_type=2,  # горизонтальная скважина
                profile_type=0,  # для горизонтальных скважин всегда 0
                type_ha=data['type_ha'],
                type_h=data['type_h'],
                results=results['table_data']
            )
            results['calculation_id'] = calculation_id
            results['input_data'] = data
            # Добавляем well_type в результаты для сессии
            results['well_type'] = 2
            
            print("Results before saving to session:", results)
            session['calculation_results'] = results
            return redirect(url_for('results'))
        else:
            # Обработка случая, когда функция расчета вернула ошибку
            # Вместо рендеринга error.html, сохраняем ошибку в results и редиректим
            results['error'] = results['error']
            session['calculation_results'] = results
            return redirect(url_for('results'))

    except ValueError as e:
        # Вместо рендеринга error.html, сохраняем ошибку в results и редиректим
        results = {'error': f"Ошибка в введенных данных: {str(e)}"}
        session['calculation_results'] = results
        return redirect(url_for('results'))
    except Exception as e:
        # Вместо рендеринга error.html, сохраняем ошибку в results и редиректим
        results = {'error': f"Ошибка при выполнении расчета: {str(e)}"}
        session['calculation_results'] = results
        return redirect(url_for('results'))

if __name__ == '__main__':
    app.run(debug=True) 